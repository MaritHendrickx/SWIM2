# -*- coding: utf-8 -*-
''' 
Automated DREAM module built in SWIM²

Creator: Marit Hendrickx (KU Leuven)
Last update: September 2025

In this script, there are several #TODO's indicated, where you can change the settings

'''

## cd C:/Users/u0139791/OneDrive - KU Leuven/Documents/GitHub/Auto-all-fields/Validation 21-22-23
## python DREAM_auto.py

import matplotlib.pyplot as plt
import matplotlib as mpl
from math import *
import numpy as np
import pandas as pd
from datetime import datetime, date
import statistics as stat
import csv
import os

import time
import tabulate

import mcmc

import pickle
from matplotlib import rc, cm

mpl.rcParams.update(mpl.rcParamsDefault)

# Color schemes: color blind & print friendly: 
CB_2_S = ['#d73027', '#fc8d59', '#fee090', '#e0f3f8', '#91bfdb', '#4575b4']
CB_2_L = ['#a50026', '#d73027', '#f46d43', '#fdae61', '#fee090', '#ffffbf', '#e0f3f8', '#abd9e9', '#74add1', '#4575b4', '#313695']
from cycler import cycler
mpl.rcParams['axes.prop_cycle'] = cycler(color=CB_2_S)

#%% Year and directory

global year
year = '2023'  #TODO

main_dir = os.path.dirname(os.path.realpath(__file__))
os.chdir(main_dir)


#%% We want to log the DREAM runs in an Excel file
import openpyxl
excel_file = 'DREAM_auto_log_'+datetime.now().strftime('%Y%m%d_%H%M')+'.xlsx'

# create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
# add headers to the worksheet
headers = ['Folder','Year','Case','Gewas','Deel teelt','Verantwoordelijke','Name','Irrigatie',
           'Observatiedata','Validatiedag','Sensorkalibratie DREAM','Sensorkalibratie','Forecast date',
           'Error/Success','Version','Run time']
for i, header in enumerate(headers, start=1):
    worksheet.cell(row=1, column=i, value=header)


#%% Loop initialisation
'''We import the logging module and set up a basic configuration 
to log errors to a file named "error.log" at the ERROR level. Within the loop, 
we catch any exceptions that occur and log the error message using the logging.error() 
function. The error message includes the case number and the exception message.
level=logging.WARNING : Error and warning levels'''

import logging
import warnings
logging.basicConfig(filename='error_warnings.log', level=logging.WARNING)
warnings.filterwarnings('ignore', category=RuntimeWarning)

#### Settings -----------------------------------------------------------------
from Final_mcmc import final_func # import shoud be at the start!
from SWB_model import settings_func
drop_samp,cases,df_sensor_teler,cal_par_on,sensor,DREAM_obsdata,lik_sigma_est,corr_est,cal,version, \
    forecast_on,date_YYYYMMDD,validation_days = settings_func(year)
    
save_extra=True # save extra figures & data # TODO

#%% Read data -----------------------------------------------------------------

df1 = pd.read_csv('soildata_212223.csv', encoding= 'unicode_escape')
df1['year'] = df1['year'].fillna(0).astype(int).astype(str)
df1['year'] = df1['year'].replace('0', '')
df1 = df1[df1['year']==year]
df1.reset_index(inplace=True,drop=True)

df2 = pd.read_csv('crop_FAO.csv', encoding= 'unicode_escape')

df_ET = pd.read_csv('ETo_212223.csv', encoding= 'unicode_escape')
df_ET = df_ET[df_ET['year'].astype(str)==year]
df_ET.reset_index(inplace=True,drop=True)

df_R = pd.read_csv('Neerslag_212223.csv', encoding= 'unicode_escape')
df_R = df_R[df_R['year'].astype(str)==year]
df_R.reset_index(inplace=True,drop=True)

df_I = pd.read_csv('Irrigatie_212223.csv', encoding= 'unicode_escape')
df_I = df_I[df_I['year'].astype(str)==year]
df_I.reset_index(inplace=True,drop=True)

df_obs=pd.read_csv('Soilobs_212223.csv', encoding= 'unicode_escape')
df_obs = df_obs[df_obs['year'].astype(str)==year]
df_obs.reset_index(inplace=True,drop=True)

df_list = [df_R,df_I,df_ET,df1,df_obs,df2]

#%% CASE SELECTION ------------------------------------------------------------
from Sensordata import sensordata as sd_func

#Choose cases #TODO

df_cases = pd.DataFrame({'years': ['2021','2022','2023','2021','2022','2023',
                                   '2021','2021','2022','2021','2022','2023',
                                   '2021','2022','2023','2023','2023','2023'],
                         'cases': ['C55D7C','C55D7C','C55D7C','C535CA','C535CA',
                                   'C535CA','C54762','2F71D1','2F71D1','012384F3720B710DEE',
                                   'C54986','C54986','C58157','C58157','C58157',
                                   'C562C6','C535C9','C54992']})

# Cases with more than 90 calibration days & >=5 samples
# df_cases = pd.DataFrame({'years': ['2022','2021','2022','2023',
#                                     '2021','2022','2023','2023'],
#                           'cases': ['C535CA','C54762','C54986','C54986',
#                                     'C58157','C58157','C58157','C562C6']})

# df_cases = pd.DataFrame({'years': ['2021','2022','2021',
#                                    '2021','2022','2021'],
#                          'cases': ['C55D7C','C55D7C','C535CA',
#                                    '2F71D1','2F71D1','012384F3720B710DEE']})

# 'C55D7C' are double cults !

cases = ['C58157'] # this can be a list of multiple cases
pluviostat = True #accurate weather data: elaborate validation (steps of 10 days) #TODO

if not isnan(validation_days): # Validating in temporal framework (at different time steps)
    if pluviostat:
        times=12 #13
    else: times = 6 # 7x met ander aantal validation_days

    # times = .. # you can also choose yourself
    
    cases = [item for item in cases for _ in range(times)]
    validation_days_original = validation_days
    

    
#%% LOOP ------------------------------------------------------------

for c,case in enumerate(cases):
    start_timer=datetime.now()
    
    os.chdir(main_dir)
    result = np.nan
    warn = False

    try:
        # run your script for the current case here
        # if there's an error, this will jump to the except block
        from SWB_model import initial_func
        ini, teler, opkomst, verantw, p1, p2, part, soil_type, crop_name, \
            irr_method, forecast, forecast_date, g0 = initial_func(case,df1,df2,
                                                                   forecast_on=forecast_on,show=True)
        loc = teler
        loc_weather = loc
        if not isnan(validation_days):
            if c>0 and times==3 and not c%times==0: # 30-60-90
                validation_days = validation_days+30 #TODO
            elif c>0 and times==2 and not c%times==0: # 55-65
                validation_days = validation_days+10 #TODO
            elif c>0 and not c%times==0:
                if pluviostat:
                    validation_days = validation_days+10 
                else: validation_days = validation_days+20  #TODO
            elif c%times==0:
                validation_days=validation_days_original
            
        os.chdir(main_dir)
        
        # if part=='0': raise NameError('We skip part 0 for now.')

        #%% Run scripts: run_mcmc
        # Set rng_seed and case study
        rng_seed=11 # reproducability
        
        CaseStudy=1
        if  CaseStudy==1: # soil water model with covariance matrix
            par=12 #parameters
            seq=6 #chains (N)
            T=4000 # uitzonderlijk meer gens nodig: 4000, otherwise 6000 #generations # beginning of season:3000
            # T=100 # test #TODO
            if year=='2022' and (case=='C54986' or case=='C535CA'): T=6000
            if lik_sigma_est: 
                T=T+2000 #TODO
                if corr_est: T=T+1000 #TODO
            ndraw=seq*T
            thin=1                  # not high-dimensional, so no thinning needed
            steps=100               # steps in iteration = steps*seq
            # jr_scale=1              # reduce scaling of jumps (jump distance) --> higher AR
            jr_scale=0.8              # 0.8 if AR<20
            #Prior='Normal'
            Prior='LHS' # noninformative prior : uniform prior with lower and upper boundaries
                
        
        ## time estimate of iterations:
        time_est=T*0.5/60
        print('Maximum estimated time:'+str(time_est)+'minutes')
        
        ## model
        from SWB_model import SWB # SWB model
        name=case
        print(name)
        
        ## parameters
        if cal_par_on:
            par_names=['Kcb_ini','Kcb_mid','Kcb_end','L_ini','L_dev','L_mid','fc','log(Ksat)','CN','GWT_max','Zr_max','v_ini','Interc a','Slope b']
        else:
            par_names=['Kcb_ini','Kcb_mid','Kcb_end','L_ini','L_dev','L_mid','fc','log(Ksat)','CN','GWT_max','Zr_max','v_ini']
            
        par=len(par_names)

        ## Initial simulation
        SWC, sw_list, g_list, sensor_data_adj, covar_adj, df_obs, _, _ = SWB(ini[0],ini[1],ini[2],ini[3], \
                               ini[4],ini[5],ini[6],ini[7],ini[8],ini[9],ini[10],ini[11], \
                               sensor=True,cal=cal, \
                               sensor_cal=np.empty(0),CI=np.empty(0),show=[False, ''],case=case,year=year, \
                                   forecast=np.empty(0),df_list=df_list)
        print(sensor_data_adj)
        
        if not isnan(validation_days):
            if len(sensor_data_adj)<validation_days-10: continue
            
        if not isnan(validation_days):
            n_samp_ori = len(df_obs)
            N_sensor_ori = len(sensor_data_adj)
            sensor_data_adj = sensor_data_adj[:validation_days]
            date = sensor_data_adj['Date'].iloc[-1]
            df_obs = df_obs.loc[df_obs['Date'] <= date]
            n_samp_val = len(df_obs)
            N_sensor_val = len(sensor_data_adj)
            
                
        ## Parameter boundaries
        
        if ini[7]>=log(200) and (soil_type=='Z' or soil_type=='S' or soil_type=='P'): # Sandy (Z,S,P)
            Ksat_LB=log(200)
            Ksat_UB=log(4000)
        elif ini[7]>=log(100) and ini[7]<=log(750) and (soil_type=='L' or soil_type=='A'): # Loamy (L,A)
            Ksat_LB=log(100)
            Ksat_UB=log(750)
        elif ini[7]>=log(5) and ini[7]<=log(150) and (soil_type=='E' or soil_type=='U'): # Clayey (E,U)
            Ksat_LB=log(5)
            Ksat_UB=log(150)
            
        if opkomst:
            LB=[0, 0.85, 0.65, ini[3]-0.5, -0.5, -0.5, min(ini[6],0.15), Ksat_LB, 58, 100, 0.3, 0.05]
            UB=[0.05, 1.15, 1.15, ini[3]+0.5, 0.5, 0.5, min(0.4,ini[6]+0.25), Ksat_UB, 91, 200, 0.9, min(0.4,ini[6]+0.25)]
        else:
            if crop_name=='witloof':
                LB=[0.05, 0.85, 0.65, ini[3]-min(ini[3],21), ini[4]-min(ini[4],7), ini[5]-min(ini[5],7), min(ini[6],0.15), Ksat_LB, 58, 100, 0.3, 0.1]
                UB=[0.4, 1.15, 1.15, ini[3] + 10, ini[4] + 7, ini[5] + 7, min(0.4,ini[6]+0.25), Ksat_UB, 91, 200, 0.9, min(0.4,ini[6]+0.25)]
            else:
                LB=[0.05, 0.85, 0.65, ini[3]-min(ini[3],10), ini[4]-min(ini[4],7), ini[5]-min(ini[5],7), min(ini[6],0.15), Ksat_LB, 58, 100, 0.3, 0.1]
                UB=[0.4, 1.15, 1.15, ini[3] + 10, ini[4] + 7, ini[5] + 7, min(0.4,ini[6]+0.25), Ksat_UB, 91, 200, 0.9, min(0.4,ini[6]+0.25)]

        if cal_par_on: 
            LB=LB+[-0.2, 0]
            UB=UB+[0.2, 3]
            
        center=(np.array(LB)+np.array(UB))/2
                
        #%% DREAM
        # Run the DREAMzs algorithm
        if __name__ == '__main__':
            
            start_time = time.time()
            ## How much chains in parallel? #TODO
            parallel_jobs=int(os.cpu_count()-2) # 6?
            parallel_jobs=seq
            q=mcmc.Sampler(df_list=df_list,data_dir=main_dir,case=case,year=year,ini=ini,LB=LB,UB=UB,opkomst=opkomst,
                            crop_name=crop_name,forecast=forecast,
                            validation_days=validation_days,
                            CaseStudy=CaseStudy,seq=seq,ndraw=ndraw,Prior=Prior,parallel_jobs=parallel_jobs,
                            steps=steps,savemodout=True,parallelUpdate = 0.9,pCR=True,thin=thin,nCR=3,DEpairs=3, #DEpairs=1
                            pJumpRate_one=0.2,BoundHandling='Reflect',lik_sigma_est=lik_sigma_est,corr_est=corr_est,DoParallel=True,
                            jr_scale=jr_scale,rng_seed=rng_seed, cal=cal,cal_par_on=cal_par_on,DREAM_obsdata=DREAM_obsdata)
            
                            # parallelUpdate = 0.9 ~ DREAMzs: 10% snooker update
                            # pCR = selection probabilities --> Adapt crossover probabilities
            
            print("Iterating ...")
            
            tmpFilePath=None # None or: main_dir+'\out_tmp.pkl'
            
            Sequences, Z, OutDiag, fx, MCMCPar, MCMCVar = q.sample(RestartFilePath=tmpFilePath)
            
            end_time = time.time()
            
            print("This sampling run took %5.4f seconds." % (end_time - start_time))
        
        
        #%% Results
    
        rc('text', usetex=False)
    
        outfilename='dreamzs_out'
        
        with open(main_dir+'/'+outfilename+'.pkl','rb') as f: # Use this to load final results only
        
            tmp_obj=pickle.load(f)
        try:
            Sequences=tmp_obj['Sequences']
            Z=tmp_obj['Z']
            OutDiag=tmp_obj['OutDiag']
            fx=tmp_obj['fx']
            MCMCPar=tmp_obj['MCMCPar']
            Measurement=tmp_obj['Measurement']
            Extra=tmp_obj['Extra']
            MCMCVar=tmp_obj['MCMCVar']
            Modelname=tmp_obj['ModelName']
            print(MCMCVar.Iter)
        except:
            pass
        del tmp_obj
        
        
        from mcmc_func import Genparset
        # The next 2 lines are for when temporary results are loaded (need to remove zeros)
        #if outfilename=='out_tmp':
        idx=np.argwhere(Sequences[:,0,0]!=0) 
        Sequences=Sequences[idx[:,0],:,:]
        
        ParSet=Genparset(Sequences)
        
        ## SAVE
        # create folder
        global folder
        folder=name+'_'+crop_name+'_'+datetime.now().strftime('%Y%m%d_%H%M')
        folder_path=os.path.join(main_dir, folder)
        os.mkdir(folder_path)
        os.chdir(folder_path)
        
        LogLik=ParSet[:,-1]
        print(LogLik)
        MaxLogLik=max(LogLik)
        print(MaxLogLik)
        print(np.where(LogLik==MaxLogLik))
        
        ParSetMax=ParSet[np.where(LogLik==MaxLogLik)[0][0]] # ParSetMax is parameter set with highest likelihood (=best guess)
        with open('ParSet_'+name+year+crop_name+'.npy', 'wb') as f:
            np.save(f, ParSet)
        with open('ParSet_MaxLL_'+name+year+crop_name+'.npy', 'wb') as f:
            np.save(f, ParSetMax)
        # if save_extra:
        #     with open('fx_'+name+year+crop_name+'.npy', 'wb') as f:
        #         np.save(f, fx)
            
        ## Validation
        if not isnan(validation_days):
            validation_df = pd.DataFrame({'Validation days':[validation_days],
                                          'N_sensor_val':[N_sensor_val],
                                          'n_samp_val':[n_samp_val],
                                          'last_date':[date]})
            validation_df.to_csv('validation_info.csv',index=False)
        
        ## Save information on measurements used in DREAM
        df_meas_DREAM = pd.DataFrame({'Validation days':[validation_days],
                                      'N_total':[Measurement.N], # total data points
                                      'N_sensor':[Measurement.n], # sensor datapoints
                                      'N_samples':[Measurement.N-Measurement.n]}) # soil samples
        df_meas_DREAM.to_csv('Measurements_used_DREAM.csv',index=False)
        df_covar_DREAM = pd.DataFrame(Measurement.Sigma) # covariance matrix
        df_covar_DREAM.to_csv('Covar_DREAM.csv',index=False)
        
        
        #%% [Read data for testing]
        # main_dir='XXX'
        # folder='C58157_witloof_20241128_1158'
        # folder_path=os.path.join(main_dir, folder)
        # os.chdir(folder_path)
        # ParSet=np.load('ParSet_C581572023witloof.npy')
        # ParSetMax=np.load('ParSet_MaxLL_C581572023witloof.npy')
        
        #%% Visualization
        ## Check acceptance rate: should be in 10% - 50% or so
        
        #if outfilename=='out_tmp':
        try:
            istart=np.int(0.8*MCMCVar.Iter/(MCMCPar.steps*MCMCPar.seq))
            iend=np.int(MCMCVar.Iter/(MCMCPar.steps*MCMCPar.seq))
        except:
            istart=np.int(0.8*MCMCPar.ndraw/(MCMCPar.steps*MCMCPar.seq))
            iend=np.int(MCMCPar.ndraw/(MCMCPar.steps*MCMCPar.seq))
        AR = np.mean(OutDiag.AR[istart:iend,1]);print('Acceptance rate (AR) = ',AR)
        plt.figure()
        plt.plot(OutDiag.AR[1:,0],OutDiag.AR[1:,1],'ob')
        plt.title('Acceptance rate (AR)')
        plt.savefig('AR_'+name+year+'.png',dpi=300)
        # plt.show()
        plt.close()
        
        #%% Check Rstat convergence 

        dummy=np.where((OutDiag.R_stat[:,1:]<=1.2) & (OutDiag.R_stat[:,1:] > 0))
        try:
            row_num, counts = np.unique(dummy[0],return_counts=True)
            row_num=row_num[np.argmax(counts==MCMCPar.n)]
            print('R_stat convergence declared at iteration '+str(OutDiag.R_stat[row_num,0]))
        except:
            print('Not converged according to R_stat')
            logging.warning(f"Warning for case {case}: Not converged according to R_stat. Redo DREAM calibration with more generations.")
            warn = True
            pass
        # Make a plot
        xx=np.copy(OutDiag.R_stat[1:iend+1,0])
        yy=np.zeros_like(xx)+1.2
        plt.figure()
        for i in range(0,MCMCPar.n):
            plt.plot(OutDiag.R_stat[1:iend+1,0],OutDiag.R_stat[1:iend+1,i+1],'-')
        plt.plot(xx,yy,'-k')
        plt.ylim(0.5,2)
        plt.title('Convergence R\u0302-statistic')
        plt.savefig('convergence_'+name+year+'.png',dpi=300)
        # plt.show()
        plt.close()
        
        #%% Save AR & convergence
        Rstat_dict = {'Rstat_'+par_names[i]: OutDiag.R_stat[1:iend+1,i+1] for i in range(par)}
        df_AR_Rstat = pd.DataFrame({'AR_steps':OutDiag.AR[1:,0],'AR':OutDiag.AR[1:,1],
                                    'Rstat_steps':OutDiag.R_stat[1:iend+1,0],
                                    **Rstat_dict})
        df_AR_Rstat.to_csv('diagnostics_AR_Rstat_'+name+year+crop_name+'.csv',index=False)

        #%% Quickly check posterior distribution using the last 50% of the samples (50% burn-in)
        istart=np.round(ParSet.shape[0]*0.5).astype('int')
        av=np.mean(ParSet[istart:,:],axis=0)    # average of posteriors (not used)
        s=np.std(ParSet[istart:,:],axis=0)    # standard deviation of posteriors (not used)
        
        plt.figure(figsize=(10,6))
        for i in range(12):
            plt.subplot(3,4,i+1).set_title(par_names[i])
            plt.xlim(LB[i],UB[i])
            plt.hist(ParSet[istart:,i],color="orange")
            line1=plt.plot(ini[i],0,'xr',mew=3,ms=10)[0] # Plot true value in red
            line1.set_clip_on(False)
            line2=plt.plot(ParSetMax[i],0,'xc',mew=3,ms=10)[0] # Plot value with max likelihood
            line2.set_clip_on(False)
            plt.ylabel('Frequency (-)')
            plt.tick_params(axis='both', pad=1)
        plt.tight_layout()
        plt.subplots_adjust(top=0.9)
        plt.suptitle('Posterior distribution', fontsize='x-large', weight='bold')
        plt.savefig('posteriors_'+name+year+'.png',dpi=300) # do here: _redo
        # plt.show()
        plt.close()

        if lik_sigma_est: # parameter 13 (index 12)
            par_names_sigma = ['log(Sigma)','Sigma']
            LB_sigma = [np.log(0.00001),0.00001]
            UB_sigma = [np.log(1),1]
            ini_sigma = [np.log(0.002068),0.002068] # pooled total sensor error variance
            ParSetMax_sigma = [ParSetMax[12],np.exp(ParSetMax[12])]
            for i in range(2):
                plt.subplot(2,1,i+1).set_title(par_names_sigma[i])
                if i==0:
                    plt.xlim(LB_sigma[i],UB_sigma[i])
                    plt.hist(ParSet[istart:,12],color="orange")
                elif i==1:
                    plt.xlim(np.exp(min(ParSet[istart:,12]))*0.9,np.exp(max(ParSet[istart:,12]))*1.1)
                    plt.hist(np.exp(ParSet[istart:,12]),color="orange")
                line1=plt.plot(ini_sigma[i],0,'xr',mew=3,ms=10)[0] # Plot true value in red
                line1.set_clip_on(False)
                line2=plt.plot(ParSetMax_sigma[i],0,'xc',mew=3,ms=10)[0] # Plot value with max likelihood
                line2.set_clip_on(False)
                plt.ylabel('Frequency (-)')
                plt.tick_params(axis='both', pad=1)
            plt.tight_layout()
            plt.subplots_adjust(top=0.9)
            plt.suptitle('Posterior distribution Sigma', fontsize='x-large', weight='bold')
            plt.savefig('posteriorsigma_'+name+year+'.png',dpi=300) # do here: _redo
            # plt.show()
            plt.close()
            
        if corr_est: # parameter 14 (index 13)
            par_names_corr = ['Error autocorrelation']
            LB_corr = [0]
            UB_corr = [1]
            ini_corr = [0.52] # pooled AR (Hendrickx et al., 2025)
            ParSetMax_corr = [ParSetMax[13]]
            
            plt.figure()
            plt.title(par_names_corr[0])
            plt.xlim(LB_corr[0],UB_corr[0])
            plt.hist(ParSet[istart:,13],color="orange")
            line1=plt.plot(ini_corr[0],0,'xr',mew=3,ms=10)[0] # Plot true value in red
            line1.set_clip_on(False)
            line2=plt.plot(ParSetMax_corr[0],0,'xc',mew=3,ms=10)[0] # Plot value with max likelihood
            line2.set_clip_on(False)
            plt.ylabel('Frequency (-)')
            plt.tick_params(axis='both', pad=1)
            plt.tight_layout()
            plt.subplots_adjust(top=0.9)
            plt.suptitle('Posterior distribution AR', fontsize='x-large', weight='bold')
            plt.savefig('posteriorAR_'+name+year+'.png',dpi=300)
            # plt.show()
            plt.close()
            
        
        #%% Check correlations using the last 50% of the samples
        
        if save_extra:
            ## Scatterplot
            par_12 = 12
            plt.figure(figsize=(8,8))
            for j in range(par_12):
                for i in range(par_12):
                    plt.subplot(par_12,par_12,1+i+j*par_12)
                    plt.ylim(LB[j],UB[j])
                    plt.xlim(LB[i],UB[i])
                    if i==j:
                        text = plt.text(center[i],center[i], par_names[i], ha="center", va="center", color="k", fontsize=8)
                        plt.tick_params(axis='both', colors='w')
                    else:
                        plt.plot(ParSet[istart:,i],ParSet[istart:,j], marker='.', linestyle='', markersize=1, color="royalblue",alpha=0.25)
                    if i>0:
                        plt.yticks([])
                    if j<par_12-1:
                        plt.xticks([])
                    plt.tick_params(axis='both', labelsize=8, pad=3)                   
            plt.tight_layout()
            plt.subplots_adjust(hspace=0.15,wspace=0.15)
            plt.savefig('scatterplot_'+name+year+'.png',dpi=300)
            # plt.show()
            plt.close()
            
            ## Correlation heatmap
            x=np.transpose(np.array(ParSet[istart:,0:par_12]))
            corr = np.corrcoef(x).round(decimals=2)
            fig, ax = plt.subplots()
            im = plt.imshow(corr,cmap="bwr")
            ax.set_xticks(np.arange(par_12))
            ax.set_yticks(np.arange(par_12))
            ax.set_xticklabels(par_names[:par_12])
            ax.set_yticklabels(par_names[:par_12])
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
                      rotation_mode="anchor")
            # Loop over data dimensions and create text annotations
            for i in range(par_12):
                for j in range(par_12):
                    text = ax.text(j, i, corr[i, j],
                                    ha="center", va="center", color="k")
            plt.colorbar(im, ax=ax, format='% .2f')
            plt.clim(-1,1)
            ax.set_title("Correlation heatmap",fontweight ="bold")
            fig.tight_layout()
            plt.savefig('correlation_'+name+year+'.png',dpi=300)
            # plt.show()
            plt.close()
        
        #%% Traceplot
        if save_extra:
            plt.figure(figsize=(9,7))
            for j in range(seq):
                for i in range(par_12):
                    plt.subplot(par_12,1,i+1)
                    plt.plot(range(T),Sequences[:,i,j],linewidth=1)
                    plt.ylabel(par_names[i], fontsize=9)
                    plt.tick_params(axis='both', labelsize=9, pad=5)
                    if i<par_12-1:
                        plt.xticks([])
            plt.xlabel('Sample of Markov Chain', fontsize=9)            
            plt.savefig('traceplot_'+name+year+'.png',dpi=300)
            # plt.show()
            plt.close()

        
        #%% Traceplot loglikelihood
        if save_extra:
            plt.figure(figsize=(9,7))
            plt.subplot(121).set_title('Traceplot loglikelihood')
            plt.plot(range(int(T*seq)),ParSet[:,-1])
            plt.ylabel("Loglikelihood")
            plt.subplot(122).set_title('Traceplot likelihood')
            plt.plot(range(int(T*seq)),np.exp(ParSet[:,-1]))
            plt.ylabel("Likelihood")
            plt.tight_layout()
            plt.savefig('loglikelihood_'+name+year+'.png',dpi=300)
            # plt.show()
            plt.close()
        
        
        #%% Calculate 95% prediction interval on the parameters using the last 50% of the samples
        
        ParSet50 = ParSet[istart:,:-2] # Parameter values excl 50% burn-in
        excl = np.round(ParSet50.shape[0]*0.025).astype('int') # excl upper and lower 2.5% to obtain 95% PI
        
        sort=ParSet50.argsort(axis=0)
        ParSet50_sort = ParSet50[sort, np.arange(sort.shape[1])] # every column is now sorted
        ParSet_95 = ParSet50_sort[excl:-excl,:] # this is the 95%
        ParSet_95_min=np.min(ParSet_95,axis=0) # lower 95% boundary
        ParSet_95_max=np.max(ParSet_95,axis=0) # upper 95% boundary
            
        # print(tabulate.tabulate([['K_ini', ParSetMax[0]," %.4f - %.4f " % (ParSet_95_min[0],ParSet_95_max[0])],
        #                           ['K_mid',ParSetMax[1], " %.4f - %.4f " % (ParSet_95_min[1],ParSet_95_max[1])],
        #                           ['K_end', ParSetMax[2], " %.4f - %.4f " % (ParSet_95_min[2],ParSet_95_max[2])],
        #                           ['L_ini', ParSetMax[3]," %.4f - %.4f " % (ParSet_95_min[3],ParSet_95_max[3])],
        #                           ['L_dev',ParSetMax[4], " %.4f - %.4f " % (ParSet_95_min[4],ParSet_95_max[4])],
        #                           ['L_mid', ParSetMax[5], " %.4f - %.4f " % (ParSet_95_min[5],ParSet_95_max[5])],
        #                           ['fc', ParSetMax[6], " %.4f - %.4f " % (ParSet_95_min[6],ParSet_95_max[6])],
        #                           ['K_sat_log', ParSetMax[7]," %.4f - %.4f " % (ParSet_95_min[7],ParSet_95_max[7])],
        #                           ['CN', ParSetMax[8]," %.4f - %.4f " % (ParSet_95_min[8],ParSet_95_max[8])],
        #                           ['Zrmax', ParSetMax[9]," %.4f - %.4f " % (ParSet_95_min[9],ParSet_95_max[9])]],
        #                         headers=['Parameter','Maximum likelihood parameter value','95% interval']))
        # if cal_par_on:
        #     print(tabulate.tabulate([['Intercept a', ParSetMax[10]," %.4f - %.4f " % (ParSet_95_min[10],ParSet_95_max[10])],
        #                               ['Slope b', ParSetMax[11]," %.4f - %.4f " % (ParSet_95_min[11],ParSet_95_max[11])]],
        #                             headers=['Parameter','Maximum likelihood parameter value','95% interval']))
        
        if lik_sigma_est:
            with open('ParSetMax-95_'+name+year+crop_name+'.csv', mode='w',newline='') as file:
                writer = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(['Parameter','Max LLH parameter value','Min 95% interval','Max 95% interval']) 
                for i in range(par):
                    try: writer.writerow([par_names[i], ParSetMax[i],ParSet_95_min[i],ParSet_95_max[i]])
                    except Exception as e:
                        print(f"Error writing row {i}: {e}")
                try:
                    writer.writerow(['log(Sigma)',ParSetMax[12],ParSet_95_min[12],ParSet_95_max[12]])
                    writer.writerow(['Sigma',np.exp(ParSetMax[12]),np.exp(ParSet_95_min[12]),np.exp(ParSet_95_max[12])])
                except Exception as e:
                    print(f"Error writing Sigma row: {e}")
        else:
            with open('ParSetMax-95_'+name+year+crop_name+'.csv', mode='w',newline='') as file:
                file = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                file.writerow(['Parameter','Max LLH parameter value','Min 95% interval','Max 95% interval']) 
                file.writerows([par_names[i], ParSetMax[i],ParSet_95_min[i],ParSet_95_max[i]] for i in range(par)) 
            
        os.chdir(main_dir)
        SWC, sw_list, g_list, sensor_data_adj, covar_adj, df_obs, _, _ = SWB(ParSetMax[0],ParSetMax[1],ParSetMax[2], \
              ParSetMax[3],ParSetMax[4],ParSetMax[5],ParSetMax[6],ParSetMax[7],ParSetMax[8],ParSetMax[9],ParSetMax[10],ParSetMax[11], \
                  sensor=True,cal=cal,sensor_cal=np.empty(0),CI=np.empty(0), \
                      show=[False,folder],case=name,year=year,forecast=forecast,df_list=df_list) # hier opslaan --> folder
        
        if cal_par_on:
            os.chdir(folder_path)
            with open('sensor_data_adj_'+name+year+crop_name+'.npy', 'wb') as f:
                np.save(f, sensor_data_adj)
            os.chdir(main_dir)
        
        
        #%% Final - separate script
        df_metrics = final_func(year,df_list,folder,'',cal_par_on,cal,forecast,simirr=False)

        #%% Log warnings
        
        if AR<10 or AR>50:
            logging.warning(f"Warning for case {case}: AR is not between 10 and 50%. Redo DREAM calibration with more generations.")
            warn = True
        if df_metrics['sensor'][0]==0:
            logging.warning(f"Warning for case {case}: No sensor data.")
            warn = True
        if df_metrics['both']['RMSE']>0.04:
            logging.warning(f"Warning for case {case}: RMSE > 0.04.")
            warn = True


        if warn: result = 'Success with warnings'
        else: result = 'Success'
        
        end_timer = datetime.now()-start_timer
        print(result, end_timer)
        
        # add a row for this case to the worksheet
        row_values = [folder, year, case, crop_name, part, verantw, teler, 
                      irr_method, DREAM_obsdata, validation_days, cal_par_on,
                      cal,forecast_date, result, version,end_timer]
        i = c+1
        for j, value in enumerate(row_values, start=1):
            worksheet.cell(row=i+1, column=j, value=value)
            
        os.chdir(main_dir)
        workbook.save(excel_file)
        
    except Exception as e:
        # if there's an error, log the message and continue to the next case
        logging.error(f"Error occurred for case {case}: {e}")
        
        end_timer = datetime.now()-start_timer
        print('Error', end_timer)
        
        try: row_values = [np.nan, year, case, crop_name, part, verantw, teler, 
                      irr_method, DREAM_obsdata, validation_days, cal_par_on,
                      cal,forecast_date, 'Error', version,end_timer]        
        except: row_values = [np.nan, year, case, np.nan, np.nan, np.nan, np.nan, 
                      np.nan, np.nan, np.nan, np.nan,
                      np.nan,np.nan, 'Error', version,end_timer]    
        
        i = c+1
        for j, value in enumerate(row_values, start=1):
            worksheet.cell(row=i+1, column=j, value=value)
        
        os.chdir(main_dir)
        workbook.save(excel_file)
    
        continue


# save the workbook to a file
os.chdir(main_dir)
workbook.save(excel_file)


#%% Read error logs
read_errors = False

if read_errors:
    import collections
    
    # read log file into a list of strings
    with open('error_warnings.log', 'r') as f:
        log_lines = f.readlines()
    
    # create dictionary to store error/warning messages and associated case numbers
    error_summary = collections.defaultdict(list)
    
    # iterate through log lines and add to summary dictionary
    for line in log_lines:
        if 'ERROR' in line or 'WARNING' in line:
            error_message = line.split(': ')[1]
            case_number = line.split(': ')[0].split()[-1]
            error_summary[error_message].append(case_number)
    
    # print summary of errors/warnings
    print('Summary of errors/warnings:')
    for error_message, cases in error_summary.items():
        print(f'{error_message} occurred in cases {", ".join(cases)}')
        
        